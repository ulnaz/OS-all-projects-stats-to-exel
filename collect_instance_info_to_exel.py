#!/usr/bin/python3 

import openpyxl
import os 
import yaml

workbook = openpyxl.load_workbook('CHD.xlsx')
sheet = workbook.active
os.popen("source /root/admin_openrc")
server_list = os.popen("openstack server list --all-projects -f value | awk '{print $1}'").read().splitlines()
for server in server_list:
    command = "openstack server show -f yaml " + server
    data = os.popen(command)
    yaml_data = yaml.load(data)
    row = server_list.index(server)
    sheet.cell(row=server_list.index(server) + 1, column=1).value = yaml_data['name']
    sheet.cell(row=server_list.index(server) + 1, column=2).value = server
    sheet.cell(row=server_list.index(server) + 1, column=3).value = yaml_data['project_id']
    sheet.cell(row=server_list.index(server) + 1, column=4).value = yaml_data['OS-EXT-SRV-ATTR:instance_name']
    sheet.cell(row=server_list.index(server) + 1, column=5).value = yaml_data['OS-EXT-SRV-ATTR:hypervisor_hostname']
    sheet.cell(row=server_list.index(server) + 1, column=6).value = yaml_data['flavor'].split(sep="(")[0].strip()
    flavor_id = yaml_data['flavor'].split(sep="(")[1][:36]
    sheet.cell(row=server_list.index(server) + 1, column=7).value = flavor_id
    command = "openstack flavor show -f yaml " + flavor_id
    flavor_data = yaml.load(os.popen(command).read())
    sheet.cell(row=server_list.index(server) + 1, column=8).value = flavor_data['ram']
    sheet.cell(row=server_list.index(server) + 1, column=9).value = flavor_data['vcpus']
    sheet.cell(row=server_list.index(server) + 1, column=10).value = flavor_data['disk']
    sheet.cell(row=server_list.index(server) + 1, column=11).value = yaml_data['addresses']
    image_name = yaml_data['image'].split(sep='(')[0].strip()
    sheet.cell(row=server_list.index(server) + 1, column=12).value = image_name
    sheet.cell(row=server_list.index(server) + 1, column=13).value = yaml_data['image'].split(sep='(')[1][:36]
    sheet.cell(row=server_list.index(server) + 1, column=14).value = yaml_data['security_groups'].replace("name=","")
    volumes_list = yaml_data['volumes_attached']
    if( not volumes_list ):
        sheet.cell(row=server_list.index(server) + 1, column=15).value = 0
    else:
        sheet.cell(row=server_list.index(server) + 1, column=15).value = volumes_list
    sheet.cell(row=server_list.index(server) + 1, column=16).value = 0
workbook.save('CHD.xlsx')
